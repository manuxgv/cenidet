# excel_model.py
import pandas as pd
from openpyxl import load_workbook

class ExcelModel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.dcc_path = 'models/BD DCC.xlsx'

#Esto se hizo el martes
    def get_all_sheets(self):
        """Obtiene los nombres de todas las hojas del archivo Excel."""
        try:
            xls = pd.ExcelFile(self.file_path)
            return xls.sheet_names
        except Exception as e:
            print(f'Error al obtener hojas del Excel: {e}')
            return []
        
    def get_sheet_data(self, sheet_name):
            """Devuelve los datos de una hoja en formato de lista de diccionarios."""
            df = self.get_sheet(sheet_name)
            return df.to_dict(orient='records') if not df.empty else []

#esto antes
    def get_sheet(self, sheet_name):
        # Cargar el archivo Excel y devolver la hoja deseada
        return pd.read_excel(self.file_path, sheet_name=sheet_name)
    

#esto viernes
    def get_sheetDCC(self, sheet_name):
        # Cargar el archivo Excel y devolver la hoja deseada
        return pd.read_excel(self.dcc_path, sheet_name=sheet_name)


    def save_data(self, sheet_name, data):
        # Cargar el archivo Excel
        wb = load_workbook(self.file_path)
        sheet = wb[sheet_name]

        # Escribir los datos en la siguiente fila disponible
        max_row = sheet.max_row + 1
        for col, value in enumerate(data, start=1):
            sheet.cell(row=max_row, column=col).value = value
        
        wb.save(self.file_path)  # Guardar el archivo

